import uuid
import re
import os
import logging
from datetime import datetime

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import DictType, ImportBatch, ImportError, ProjectApproval, ProjectMaster, ProjectMasterLog, UserProfile


DICT_CODES = [
    "ORG",
    "BUSINESS_UNIT",
    "DEPT",
    "PROJECT_TYPE",
    "ORG_MODE",
    "DATA_STATUS",
    "PROVINCE",
    "CITY",
]

PJ_CODE_PATTERN = re.compile(r"^PJ\d{10}$")
logger = logging.getLogger(__name__)

PERMISSION_FIELDS = [
    ("can_user_manage", "用户管理"),
    ("can_create_project", "新增项目"),
    ("can_query_project", "查询项目"),
    ("can_update_project", "信息更新"),
    ("can_view_project_list", "项目列表"),
    ("can_approval_manage", "审批管理"),
]


def _get_user_permissions(user):
    profile, _ = UserProfile.objects.get_or_create(user=user)
    permissions = {
        key: bool(getattr(profile, key, False))
        for key, _ in PERMISSION_FIELDS
    }
    if user.is_superuser:
        for key, _ in PERMISSION_FIELDS:
            permissions[key] = True
    return permissions


def _redirect_no_permission(request):
    messages.error(request, "无权限")
    return redirect(f"{reverse('project_master_list')}?show_add=1")


def _normalize_project_code(value):
    return re.sub(r"\s+", "", str(value or "")).upper()


def _load_dicts():
    types = (
        DictType.objects.filter(code__in=DICT_CODES, is_active=True)
        .prefetch_related("items")
        .order_by("code")
    )
    dicts = {code: [] for code in DICT_CODES}
    for dt in types:
        dicts[dt.code] = list(
            dt.items.filter(is_active=True).order_by("sort_order", "code")
        )
    return dicts


def _dict_name_map(dicts):
    return {code: {item.code: item.name for item in items} for code, items in dicts.items()}


def home(request):
    if request.user.is_authenticated:
        return redirect("project_master_list")

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        remember = request.POST.get("remember")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if not remember:
                request.session.set_expiry(0)
            return redirect("project_master_list")
        messages.error(request, "用户名或密码错误")

    return render(request, "home.html")


def logout_view(request):
    logout(request)
    return redirect("home")


@login_required
def project_master_list(request):
    permissions = _get_user_permissions(request.user)
    dicts = _load_dicts()
    dict_map = _dict_name_map(dicts)
    is_list_filter = request.GET.get("list_filter") == "1"
    force_add_form = request.GET.get("show_add") == "1"
    query_keys = [
        "project_code",
        "project_name",
        "org_name",
        "parent_pj_code",
        "province_code",
        "business_unit",
        "dept",
        "project_type",
        "org_mode",
        "data_status",
        "is_execution_level",
        "project_year",
        "created_by",
        "remark",
    ]
    has_query_request = any(request.GET.get(key, "").strip() for key in query_keys)

    if request.method == "GET":
        if is_list_filter and not permissions["can_view_project_list"]:
            return _redirect_no_permission(request)
        if has_query_request and not permissions["can_query_project"]:
            return _redirect_no_permission(request)
        if (
            not is_list_filter
            and not has_query_request
            and not force_add_form
            and not permissions["can_view_project_list"]
        ):
            return _redirect_no_permission(request)

    if request.method == "POST" and request.POST.get("form_type") == "create":
        if not permissions["can_create_project"]:
            return _redirect_no_permission(request)
        project_data = {
            "project_code": request.POST.get("project_code", "").strip(),
            "project_name": request.POST.get("project_name", "").strip(),
            "org_name": request.POST.get("org_name", "").strip(),
            "parent_pj_code": request.POST.get("parent_pj_code", "").strip() or None,
            "province_code": request.POST.get("province_code", "").strip(),
            "city_code": request.POST.get("city_code", "").strip(),
            "business_unit": request.POST.get("business_unit", "").strip(),
            "dept": request.POST.get("dept", "").strip(),
            "project_type": request.POST.get("project_type", "").strip(),
            "org_mode": request.POST.get("org_mode", "").strip(),
            "data_status": request.POST.get("data_status", "").strip(),
            "is_execution_level": request.POST.get("is_execution_level", "false")
            == "true",
            "created_by": request.POST.get("created_by", "").strip()
            or request.user.username,
            "updated_by": request.user.username,
            "remark": request.POST.get("remark", "").strip(),
        }

        if not project_data["city_code"]:
            project_data["city_code"] = project_data["province_code"]

        if project_data["project_code"] and len(project_data["project_code"]) >= 6:
            project_data["project_year"] = project_data["project_code"][2:6]
        else:
            project_data["project_year"] = ""

        try:
            with transaction.atomic():
                project = ProjectMaster(**project_data)
                project.full_clean()
                project.save()
                ProjectMasterLog.objects.create(
                    project_code=project.project_code,
                    action="create",
                    after_data=project_data,
                    operator=request.user.username,
                    source="web",
                )
            messages.success(request, "项目已新增")
            return redirect("project_master_list")
        except ValidationError as exc:
            messages.error(request, f"保存失败：{exc}")
        except Exception as exc:
            messages.error(request, f"保存失败：{exc}")

    show_update_panel = False

    if request.method == "POST" and request.POST.get("form_type") == "update":
        if not permissions["can_update_project"]:
            return _redirect_no_permission(request)
        target_code = request.POST.get("update_project_code", "").strip()
        project = ProjectMaster.objects.filter(project_code=target_code, is_deleted=False).first()
        if not project:
            messages.error(request, "未找到需要更新的项目")
            return redirect("project_master_list")

        before = {
            "project_name": project.project_name,
            "org_name": project.org_name,
            "parent_pj_code": project.parent_pj_code,
            "province_code": project.province_code,
            "business_unit": project.business_unit,
            "dept": project.dept,
            "project_type": project.project_type,
            "org_mode": project.org_mode,
            "data_status": project.data_status,
            "is_execution_level": project.is_execution_level,
            "remark": project.remark,
        }

        project.project_name = request.POST.get("project_name", project.project_name).strip()
        project.org_name = request.POST.get("org_name", project.org_name).strip()
        project.parent_pj_code = request.POST.get("parent_pj_code", "").strip() or None
        project.province_code = request.POST.get("province_code", project.province_code).strip()
        if not project.city_code:
            project.city_code = project.province_code
        project.business_unit = request.POST.get("business_unit", project.business_unit).strip()
        project.dept = request.POST.get("dept", project.dept).strip()
        project.project_type = request.POST.get("project_type", project.project_type).strip()
        project.org_mode = request.POST.get("org_mode", project.org_mode).strip()
        project.data_status = request.POST.get("data_status", project.data_status).strip()
        project.is_execution_level = request.POST.get("is_execution_level", "false") == "true"
        project.remark = request.POST.get("remark", "").strip()
        project.updated_by = request.user.username
        change_note = request.POST.get("update_note", "").strip()

        extra_keys = request.POST.getlist("update_field_key")
        extra_vals = request.POST.getlist("update_field_value")
        for key, val in zip(extra_keys, extra_vals):
            if not key:
                continue
            val = (val or "").strip()
            if val == "":
                continue
            if key == "is_execution_level":
                project.is_execution_level = val in ["true", "True", "是", "1"]
            elif key == "parent_pj_code":
                project.parent_pj_code = val or None
            elif key == "province_code":
                project.province_code = val
                if not project.city_code:
                    project.city_code = val
            elif key == "remark":
                project.remark = val
            elif hasattr(project, key):
                setattr(project, key, val)

        try:
            with transaction.atomic():
                project.full_clean()
                project.save()
                ProjectMasterLog.objects.create(
                    project_code=project.project_code,
                    action="update",
                    before_data=before,
                    after_data={
                        "project_name": project.project_name,
                        "org_name": project.org_name,
                        "parent_pj_code": project.parent_pj_code,
                        "province_code": project.province_code,
                        "business_unit": project.business_unit,
                        "dept": project.dept,
                        "project_type": project.project_type,
                        "org_mode": project.org_mode,
                        "data_status": project.data_status,
                        "is_execution_level": project.is_execution_level,
                        "remark": project.remark,
                    },
                    change_note=change_note,
                    operator=request.user.username,
                    source="update-panel",
                )
            messages.success(request, "更新成功")
        except ValidationError as exc:
            messages.error(request, f"更新失败：{exc}")
        except Exception as exc:
            messages.error(request, f"更新失败：{exc}")
        show_update_panel = True

    qs = ProjectMaster.objects.filter(is_deleted=False)
    if not permissions["can_view_project_list"] and not (has_query_request and permissions["can_query_project"]):
        qs = ProjectMaster.objects.none()
    
    if is_list_filter:
        # 项目列表独立筛选
        list_filter = {
            "project_code": request.GET.get("list_project_code", "").strip(),
            "project_name": request.GET.get("list_project_name", "").strip(),
            "org_name": request.GET.get("list_org_name", "").strip(),
            "province_code": request.GET.get("list_province_code", "").strip(),
            "business_unit": request.GET.get("list_business_unit", "").strip(),
            "dept": request.GET.get("list_dept", "").strip(),
            "data_status": request.GET.get("list_data_status", "").strip(),
            "project_year": request.GET.get("list_project_year", "").strip(),
            "created_by": request.GET.get("list_created_by", "").strip(),
        }
        
        if list_filter["project_code"]:
            qs = qs.filter(project_code__icontains=list_filter["project_code"])
        if list_filter["project_name"]:
            qs = qs.filter(project_name__icontains=list_filter["project_name"])
        if list_filter["org_name"]:
            qs = qs.filter(org_name__icontains=list_filter["org_name"])
        if list_filter["province_code"]:
            qs = qs.filter(province_code=list_filter["province_code"])
        if list_filter["business_unit"]:
            qs = qs.filter(business_unit=list_filter["business_unit"])
        if list_filter["dept"]:
            qs = qs.filter(dept=list_filter["dept"])
        if list_filter["data_status"]:
            qs = qs.filter(data_status=list_filter["data_status"])
        if list_filter["project_year"]:
            qs = qs.filter(project_year__icontains=list_filter["project_year"])
        if list_filter["created_by"]:
            qs = qs.filter(created_by__icontains=list_filter["created_by"])
        
        search = {
            "project_code": "",
            "project_name": "",
            "org_name": "",
            "parent_pj_code": "",
            "province_code": "",
            "business_unit": "",
            "dept": "",
            "project_type": "",
            "org_mode": "",
            "data_status": "",
            "is_execution_level": "",
            "project_year": "",
            "created_by": "",
            "remark": "",
        }
    else:
        # 查询项目筛选
        search = {
            "project_code": request.GET.get("project_code", "").strip(),
            "project_name": request.GET.get("project_name", "").strip(),
            "org_name": request.GET.get("org_name", "").strip(),
            "parent_pj_code": request.GET.get("parent_pj_code", "").strip(),
            "province_code": request.GET.get("province_code", "").strip(),
            "business_unit": request.GET.get("business_unit", "").strip(),
            "dept": request.GET.get("dept", "").strip(),
            "project_type": request.GET.get("project_type", "").strip(),
            "org_mode": request.GET.get("org_mode", "").strip(),
            "data_status": request.GET.get("data_status", "").strip(),
            "is_execution_level": request.GET.get("is_execution_level", "").strip(),
            "project_year": request.GET.get("project_year", "").strip(),
            "created_by": request.GET.get("created_by", "").strip(),
            "remark": request.GET.get("remark", "").strip(),
        }

        if search["project_code"]:
            qs = qs.filter(project_code__icontains=search["project_code"])
        if search["project_name"]:
            qs = qs.filter(project_name__icontains=search["project_name"])
        if search["org_name"]:
            qs = qs.filter(org_name__icontains=search["org_name"])
        if search["parent_pj_code"]:
            qs = qs.filter(parent_pj_code__icontains=search["parent_pj_code"])
        if search["province_code"]:
            qs = qs.filter(province_code=search["province_code"])
        if search["business_unit"]:
            qs = qs.filter(business_unit=search["business_unit"])
        if search["dept"]:
            qs = qs.filter(dept=search["dept"])
        if search["project_type"]:
            qs = qs.filter(project_type=search["project_type"])
        if search["org_mode"]:
            qs = qs.filter(org_mode=search["org_mode"])
        if search["data_status"]:
            qs = qs.filter(data_status=search["data_status"])
        if search["is_execution_level"] == "true":
            qs = qs.filter(is_execution_level=True)
        if search["is_execution_level"] == "false":
            qs = qs.filter(is_execution_level=False)
        if search["project_year"]:
            qs = qs.filter(project_year__icontains=search["project_year"])
        if search["created_by"]:
            qs = qs.filter(created_by__icontains=search["created_by"])
        if search["remark"]:
            qs = qs.filter(remark__icontains=search["remark"])
        
        list_filter = {
            "project_code": "",
            "project_name": "",
            "org_name": "",
            "province_code": "",
            "business_unit": "",
            "dept": "",
            "data_status": "",
            "project_year": "",
            "created_by": "",
        }

    projects = list(qs.order_by("-project_code"))
    name_map = _dict_name_map(dicts)
    for project in projects:
        project.province_name = name_map.get("PROVINCE", {}).get(
            project.province_code, project.province_code
        )
        project.city_name = name_map.get("CITY", {}).get(project.city_code, project.city_code)
        project.business_unit_name = name_map.get("BUSINESS_UNIT", {}).get(
            project.business_unit, project.business_unit
        )
        project.dept_name = name_map.get("DEPT", {}).get(project.dept, project.dept)
        project.org_mode_name = name_map.get("ORG_MODE", {}).get(project.org_mode, project.org_mode)
        project.data_status_name = name_map.get("DATA_STATUS", {}).get(project.data_status, project.data_status)
        project.project_type_name = name_map.get("PROJECT_TYPE", {}).get(
            project.project_type, project.project_type
        )

    latest_errors = []
    latest_batch = ImportBatch.objects.order_by("-imported_at").first()
    if latest_batch:
        latest_errors = list(latest_batch.errors.all())

    recent_updates = list(
        ProjectMasterLog.objects.filter(action="update").order_by("-created_at")[:10]
    )
    if not permissions["can_update_project"]:
        recent_updates = []
    project_name_map = dict(
        ProjectMaster.objects.filter(is_deleted=False).values_list("project_code", "project_name")
    )
    update_dict_key_map = {
        "province_code": "PROVINCE",
        "business_unit": "BUSINESS_UNIT",
        "dept": "DEPT",
        "project_type": "PROJECT_TYPE",
        "org_mode": "ORG_MODE",
        "data_status": "DATA_STATUS",
    }

    def _display_update_value(field_key, raw_value):
        if raw_value in (None, ""):
            return "-"

        if field_key == "is_execution_level":
            if isinstance(raw_value, bool):
                return "是" if raw_value else "否"
            return "是" if str(raw_value).strip().lower() in {"true", "1", "是"} else "否"

        if field_key == "parent_pj_code":
            code = str(raw_value)
            name = project_name_map.get(code)
            return f"{code} - {name}" if name else code

        dict_code = update_dict_key_map.get(field_key)
        if dict_code:
            code = str(raw_value)
            name = name_map.get(dict_code, {}).get(code)
            return f"{code} - {name}" if name else code

        return str(raw_value)

    field_labels = {
        "project_name": "项目名称",
        "org_name": "项目机构",
        "parent_pj_code": "上级PJ编码",
        "province_code": "所在省",
        "business_unit": "业务板块",
        "dept": "项目承担部门",
        "project_type": "项目类型",
        "org_mode": "项目组织模式",
        "data_status": "主数据系统数据状态",
        "is_execution_level": "是否为执行层",
        "remark": "备注",
    }
    for log in recent_updates:
        before = log.before_data or {}
        after = log.after_data or {}
        changed = []
        before_lines = []
        after_lines = []
        for key, label in field_labels.items():
            before_val = before.get(key)
            after_val = after.get(key)
            if before_val != after_val:
                changed.append(label)
                before_lines.append(f"{label}：{_display_update_value(key, before_val)}")
                after_lines.append(f"{label}：{_display_update_value(key, after_val)}")
        log.changed_fields = "、".join(changed) if changed else "无"
        log.before_summary = "\n".join(before_lines) if before_lines else "无"
        log.after_summary = "\n".join(after_lines) if after_lines else "无"

    update_code = request.GET.get("update_code", "").strip()
    
    # 获取所有项目用于下拉选择
    all_projects = []
    if permissions["can_update_project"]:
        all_projects = list(ProjectMaster.objects.filter(is_deleted=False).order_by("-project_code")[:500])
    
    # 获取审批数据
    from .models import ProjectApproval
    pending_approvals = []
    processed_approvals = []
    if permissions["can_approval_manage"]:
        pending_approvals = ProjectApproval.objects.filter(status="pending").order_by("-submit_time")
        processed_approvals = ProjectApproval.objects.filter(
            status__in=["approved", "rejected"]
        ).order_by("-approve_time")[:50]

    update_target = None
    if update_code and permissions["can_update_project"]:
        update_target = ProjectMaster.objects.filter(
            project_code=update_code, is_deleted=False
        ).first()
        show_update_panel = True

    # 检查是否有查询参数，如果有则显示查询表单
    has_search_params = any([
        search["project_code"],
        search["project_name"],
        search["org_name"],
        search["parent_pj_code"],
        search["province_code"],
        search["business_unit"],
        search["dept"],
        search["project_type"],
        search["org_mode"],
        search["data_status"],
        search["is_execution_level"],
        search["project_year"],
        search["created_by"],
        search["remark"],
    ])
    is_query_result = (not is_list_filter) and has_search_params
    show_action_column = (not is_query_result) and (
        permissions["can_update_project"] or permissions["can_approval_manage"]
    )

    # 检查是否是新增项目表单提交后的显示
    show_add_form = (request.method == "POST" and request.POST.get("form_type") == "create") or force_add_form
    if not permissions["can_create_project"]:
        show_add_form = False
    
    # 构建项目列表导出参数
    list_export_params = "&".join([f"{k}={v}" for k, v in list_filter.items() if v])
    if list_export_params:
        list_export_params = "list_filter=1&" + list_export_params
    else:
        list_export_params = "list_filter=1"
    
    # 获取项目年份列表
    project_years = sorted(
        set(
            ProjectMaster.objects.filter(is_deleted=False)
            .exclude(project_year="")
            .values_list("project_year", flat=True)
        ),
        reverse=True,
    )

    return render(
        request,
        "project_master_list.html",
        {
            "projects": projects,
            "dicts": dicts,
            "latest_errors": latest_errors,
            "search": search,
            "list_filter": list_filter,
            "is_list_filter": is_list_filter,
            "list_export_params": list_export_params,
            "project_years": project_years,
            "recent_updates": recent_updates,
            "update_target": update_target,
            "update_code": update_code,
            "show_update_panel": show_update_panel,
            "show_search_form": has_search_params and permissions["can_query_project"],
            "is_query_result": is_query_result,
            "show_action_column": show_action_column,
            "show_add_form": show_add_form,
            "update_now": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "all_projects": all_projects,
            "pending_approvals": pending_approvals,
            "processed_approvals": processed_approvals,
            "permissions": permissions,
        },
    )


@login_required
def export_project_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "项目主数据模板"
    headers = [
        "项目主数据编码",
        "项目名称",
        "项目机构名称",
        "上级PJ编码",
        "所在省",
        "业务板块",
        "项目承担部门",
        "项目类型",
        "项目组织模式",
        "主数据系统数据状态",
        "是否为执行层",
        "备注",
    ]
    ws.append(headers)

    # 添加第二个子表：字典参考
    ws_dict = wb.create_sheet(title="字典参考")
    dict_headers = ["字典类型", "编码", "名称", "说明"]
    ws_dict.append(dict_headers)

    # 加载字典数据
    dicts = _load_dicts()
    dict_types_map = {
        "PROVINCE": "所在省",
        "BUSINESS_UNIT": "业务板块",
        "DEPT": "项目承担部门",
        "PROJECT_TYPE": "项目类型",
        "ORG_MODE": "项目组织模式",
        "DATA_STATUS": "数据状态",
    }

    for dict_type_code, type_name in dict_types_map.items():
        items = dicts.get(dict_type_code, [])
        for item in items:
            ws_dict.append([
                type_name,
                item.code,
                item.name,
                item.remark or ""
            ])

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40

    ws_dict.column_dimensions['A'].width = 15
    ws_dict.column_dimensions['B'].width = 15
    ws_dict.column_dimensions['C'].width = 50
    ws_dict.column_dimensions['D'].width = 30

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=project_master_template.xlsx"
    wb.save(response)
    return response


@login_required
def export_project_list(request):
    """导出项目列表（支持筛选条件）"""
    permissions = _get_user_permissions(request.user)
    if not permissions["can_view_project_list"]:
        return _redirect_no_permission(request)

    qs = ProjectMaster.objects.filter(is_deleted=False)
    
    # 应用筛选条件
    if request.GET.get("list_filter") == "1":
        if request.GET.get("list_project_code"):
            qs = qs.filter(project_code__icontains=request.GET.get("list_project_code").strip())
        if request.GET.get("list_project_name"):
            qs = qs.filter(project_name__icontains=request.GET.get("list_project_name").strip())
        if request.GET.get("list_org_name"):
            qs = qs.filter(org_name__icontains=request.GET.get("list_org_name").strip())
        if request.GET.get("list_province_code"):
            qs = qs.filter(province_code=request.GET.get("list_province_code").strip())
        if request.GET.get("list_business_unit"):
            qs = qs.filter(business_unit=request.GET.get("list_business_unit").strip())
        if request.GET.get("list_dept"):
            qs = qs.filter(dept=request.GET.get("list_dept").strip())
        if request.GET.get("list_data_status"):
            qs = qs.filter(data_status=request.GET.get("list_data_status").strip())
        if request.GET.get("list_project_year"):
            qs = qs.filter(project_year__icontains=request.GET.get("list_project_year").strip())
        if request.GET.get("list_created_by"):
            qs = qs.filter(created_by__icontains=request.GET.get("list_created_by").strip())
    
    dicts = _load_dicts()
    name_map = _dict_name_map(dicts)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "项目列表"
    headers = [
        "项目编码",
        "项目名称",
        "项目机构名称",
        "上级PJ编码",
        "所在省",
        "业务板块",
        "承担部门",
        "项目类型",
        "组织模式",
        "数据状态",
        "是否执行层",
        "项目年份",
        "创建人",
        "创建时间",
        "备注",
    ]
    ws.append(headers)
    
    for project in qs.order_by("-created_at"):
        ws.append([
            project.project_code,
            project.project_name,
            project.org_name,
            project.parent_pj_code or "",
            name_map.get("PROVINCE", {}).get(project.province_code, project.province_code),
            name_map.get("BUSINESS_UNIT", {}).get(project.business_unit, project.business_unit or ""),
            name_map.get("DEPT", {}).get(project.dept, project.dept or ""),
            name_map.get("PROJECT_TYPE", {}).get(project.project_type, project.project_type or ""),
            name_map.get("ORG_MODE", {}).get(project.org_mode, project.org_mode or ""),
            name_map.get("DATA_STATUS", {}).get(project.data_status, project.data_status or ""),
            "是" if project.is_execution_level else "否",
            project.project_year or "",
            project.created_by or "",
            project.created_at.strftime("%Y-%m-%d %H:%M") if project.created_at else "",
            project.remark or "",
        ])
    
    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=project_list_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(response)
    return response


@login_required
def import_project_master(request):
    permissions = _get_user_permissions(request.user)
    if not permissions["can_approval_manage"]:
        return _redirect_no_permission(request)

    if request.method != "POST":
        return redirect("project_master_list")

    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "请选择要导入的文件")
        return redirect("project_master_list")

    # 保存上传的文件到临时位置
    import tempfile
    import os
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        for chunk in file.chunks():
            tmp_file.write(chunk)
        tmp_file_path = tmp_file.name
    
    try:
        # 验证文件内容
        wb = load_workbook(tmp_file_path, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        # 检查必要的列是否存在
        required_columns = ["项目主数据编码", "项目名称"]
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            os.unlink(tmp_file_path)
            messages.error(request, f"导入文件缺少必要的列: {', '.join(missing_columns)}")
            return redirect("project_master_list")
        
        # 先校验项目编码格式，避免后续数据库长度/格式异常
        code_idx = headers.index("项目主数据编码")
        invalid_rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_code = row[code_idx] if code_idx < len(row) else ""
            code = _normalize_project_code(raw_code)
            if not code:
                continue
            if not PJ_CODE_PATTERN.fullmatch(code):
                invalid_rows.append(f"第{row_num}行:{code}")
                if len(invalid_rows) >= 5:
                    break

        if invalid_rows:
            os.unlink(tmp_file_path)
            messages.error(
                request,
                "项目主数据编码格式错误，需为PJ开头+10位数字。" + "；".join(invalid_rows),
            )
            return redirect("project_master_list")

        # 统计导入数据条数
        total_rows = ws.max_row - 1  # 减去表头
        
        # 创建导入审批记录
        from .models import ProjectApproval
        approval = ProjectApproval.objects.create(
            project_code=f"IM{uuid.uuid4().hex[:10].upper()}",
            project_name=f"批量导入 {total_rows} 条项目数据",
            approval_type="import",
            submitter=request.user.username,
            change_note=f"导入文件: {file.name}, 预计导入 {total_rows} 条数据",
            import_file_path=tmp_file_path,
        )
        
        messages.success(request, f"导入申请已提交，共 {total_rows} 条数据，等待倪明珠审批。")
        return redirect("project_master_list")
        
    except Exception as exc:
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)
        messages.error(request, f"导入文件验证失败: {str(exc)}")
        return redirect("project_master_list")


@login_required
def project_master_edit(request, project_code):
    permissions = _get_user_permissions(request.user)
    if not permissions["can_update_project"]:
        return _redirect_no_permission(request)

    dicts = _load_dicts()
    project = get_object_or_404(ProjectMaster, project_code=project_code, is_deleted=False)

    if request.method == "POST":
        existing = ProjectApproval.objects.filter(
            project_code=project_code,
            approval_type="update",
            status="pending",
        ).first()
        if existing:
            messages.warning(request, "该项目已有待审批的修改申请，请勿重复提交")
            return redirect("project_master_list")

        before = {
            "project_name": project.project_name,
            "org_name": project.org_name,
            "parent_pj_code": project.parent_pj_code,
            "province_code": project.province_code,
            "city_code": project.city_code,
            "business_unit": project.business_unit,
            "dept": project.dept,
            "project_type": project.project_type,
            "org_mode": project.org_mode,
            "data_status": project.data_status,
            "is_execution_level": project.is_execution_level,
            "status": project.status,
            "remark": project.remark,
        }

        after_data = {
            "project_name": request.POST.get("project_name", "").strip(),
            "org_name": request.POST.get("org_name", "").strip(),
            "parent_pj_code": request.POST.get("parent_pj_code", "").strip() or None,
            "province_code": request.POST.get("province_code", "").strip(),
            "city_code": request.POST.get("city_code", "").strip() or request.POST.get("province_code", "").strip(),
            "business_unit": request.POST.get("business_unit", "").strip(),
            "dept": request.POST.get("dept", "").strip(),
            "project_type": request.POST.get("project_type", "").strip(),
            "org_mode": request.POST.get("org_mode", "").strip(),
            "data_status": request.POST.get("data_status", "").strip(),
            "is_execution_level": request.POST.get("is_execution_level", "false") == "true",
            "status": request.POST.get("status", "").strip() or project.status,
            "remark": request.POST.get("remark", "").strip(),
        }

        field_labels = {
            "project_name": "项目名称",
            "org_name": "项目机构",
            "parent_pj_code": "上级PJ编码",
            "province_code": "所在省",
            "city_code": "所在市",
            "business_unit": "业务板块",
            "dept": "项目承担部门",
            "project_type": "项目类型",
            "org_mode": "项目组织模式",
            "data_status": "主数据系统数据状态",
            "is_execution_level": "是否为执行层",
            "status": "状态",
            "remark": "备注",
        }

        changed_fields = [
            field_labels[key]
            for key in after_data.keys()
            if before.get(key) != after_data.get(key)
        ]

        if not changed_fields:
            messages.info(request, "未检测到字段变化，无需提交审批")
            return redirect("project_master_list")

        try:
            for key, value in after_data.items():
                setattr(project, key, value)
            project.updated_by = request.user.username
            project.full_clean()

            approval = ProjectApproval.objects.create(
                project_code=project.project_code,
                project_name=after_data.get("project_name") or project.project_name,
                approval_type="update",
                before_data=before,
                after_data=after_data,
                change_note="变更字段：" + "、".join(changed_fields),
                submitter=request.user.username,
                approver="倪明珠",
                status="pending",
            )
            messages.success(request, f"修改申请已提交，等待倪明珠审批。审批单号：{approval.id}")
            return redirect("project_master_list")
        except ValidationError as exc:
            messages.error(request, f"提交审批失败：{exc}")
        except Exception as exc:
            messages.error(request, f"提交审批失败：{exc}")

    latest_update = (
        ProjectMasterLog.objects.filter(project_code=project.project_code, action="update")
        .order_by("-created_at")
        .first()
    )

    return render(
        request,
        "project_master_edit.html",
        {
            "project": project,
            "dicts": dicts,
            "latest_update": latest_update,
        },
    )


@login_required
@login_required
def submit_delete_approval(request):
    """提交删除审批申请"""
    permissions = _get_user_permissions(request.user)
    if not permissions["can_approval_manage"]:
        return _redirect_no_permission(request)

    if request.method != "POST":
        return redirect("project_master_list")

    project_code = request.POST.get("project_code", "").strip()
    change_note = request.POST.get("change_note", "").strip()

    project = get_object_or_404(ProjectMaster, project_code=project_code, is_deleted=False)

    # 检查是否已有待审批的删除申请
    existing = ProjectApproval.objects.filter(
        project_code=project_code,
        approval_type="delete",
        status="pending"
    ).first()

    if existing:
        messages.warning(request, f"项目 {project_code} 已有一个待审批的删除申请，请勿重复提交")
        return redirect("project_master_list")

    # 创建审批记录
    before_data = {
        "project_code": project.project_code,
        "project_name": project.project_name,
        "org_name": project.org_name,
        "data_status": project.data_status,
    }

    approval = ProjectApproval.objects.create(
        project_code=project_code,
        approval_type="delete",
        before_data=before_data,
        change_note=change_note,
        submitter=request.user.username,
        approver="倪明珠",
        status="pending",
    )

    messages.success(request, f"删除申请已提交，等待倪明珠审批。审批单号：{approval.id}")
    return redirect("project_master_list")


@login_required
def approval_list(request):
    """审批列表页面"""
    # 审批管理已集成在项目列表页，避免缺失模板导致500
    return redirect("project_master_list")


@login_required
def approve_action(request, approval_id):
    """处理审批"""
    permissions = _get_user_permissions(request.user)
    if not permissions["can_approval_manage"]:
        return _redirect_no_permission(request)

    if request.method != "POST":
        return redirect("project_master_list")

    approval = get_object_or_404(ProjectApproval, id=approval_id, status="pending")

    # 检查当前用户是否为审批人
    if request.user.username != approval.approver:
        messages.error(request, "您没有权限审批此申请")
        return redirect("project_master_list")

    action = request.POST.get("action")
    approve_note = request.POST.get("approve_note", "").strip()

    if action not in {"approve", "reject"}:
        messages.error(request, "无效的审批动作")
        return redirect("project_master_list")

    try:
        with transaction.atomic():
            if action == "approve":
                # 执行修改审批
                if approval.approval_type == "update":
                    project = ProjectMaster.objects.filter(
                        project_code=approval.project_code,
                        is_deleted=False,
                    ).first()

                    if not project:
                        messages.error(request, "目标项目不存在或已删除，无法审批通过")
                        return redirect("project_master_list")

                    after_data = approval.after_data or {}
                    editable_fields = [
                        "project_name",
                        "org_name",
                        "parent_pj_code",
                        "province_code",
                        "city_code",
                        "business_unit",
                        "dept",
                        "project_type",
                        "org_mode",
                        "data_status",
                        "is_execution_level",
                        "status",
                        "remark",
                    ]
                    for key in editable_fields:
                        if key in after_data:
                            value = after_data[key]
                            if key == "is_execution_level" and isinstance(value, str):
                                value = value.strip().lower() in {"true", "1", "是"}
                            setattr(project, key, value)
                    project.updated_by = approval.submitter
                    project.full_clean()
                    project.save()

                    ProjectMasterLog.objects.create(
                        project_code=project.project_code,
                        action="update",
                        before_data=approval.before_data,
                        after_data=approval.after_data,
                        operator=approval.submitter,
                        source="approval",
                    )

                # 执行实际的删除操作
                elif approval.approval_type == "delete":
                    project = ProjectMaster.objects.filter(
                        project_code=approval.project_code,
                        is_deleted=False
                    ).first()

                    if project:
                        project.is_deleted = True
                        project.updated_by = approval.submitter
                        project.save(update_fields=["is_deleted", "updated_by"])

                        ProjectMasterLog.objects.create(
                            project_code=project.project_code,
                            action="delete",
                            before_data=approval.before_data,
                            operator=approval.submitter,
                            source="approval",
                        )
                
                # 执行导入操作
                elif approval.approval_type == "import":
                    if approval.import_file_path and os.path.exists(approval.import_file_path):
                        _process_import_file(approval.import_file_path, approval.submitter)
                        messages.success(request, "导入文件已处理完成")
                    else:
                        messages.error(request, "导入文件不存在或已过期")
                        return redirect("project_master_list")

                approval.status = "approved"
                approval.approve_time = timezone.now()
                approval.approve_note = approve_note
                approval.save(update_fields=["status", "approve_time", "approve_note"])

                messages.success(request, f"已批准 {approval.submitter} 的{approval.get_approval_type_display()}申请")

            elif action == "reject":
                # 如果拒绝导入申请，删除临时文件
                if approval.approval_type == "import" and approval.import_file_path:
                    if os.path.exists(approval.import_file_path):
                        try:
                            os.unlink(approval.import_file_path)
                        except OSError:
                            logger.warning("删除导入临时文件失败: %s", approval.import_file_path, exc_info=True)
                
                approval.status = "rejected"
                approval.approve_time = timezone.now()
                approval.approve_note = approve_note
                approval.save(update_fields=["status", "approve_time", "approve_note"])

                messages.success(request, f"已拒绝 {approval.submitter} 的{approval.get_approval_type_display()}申请")
    except Exception as exc:
        logger.exception("审批处理失败 approval_id=%s action=%s", approval_id, action)
        messages.error(request, f"审批处理失败: {exc}")
        return redirect("project_master_list")

    return redirect("project_master_list")


def _process_import_file(file_path, submitter):
    """处理导入文件"""
    from openpyxl import load_workbook
    
    dicts = _load_dicts()
    name_map = _dict_name_map(dicts)
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    header_map = {
        "项目主数据编码": "project_code",
        "项目名称": "project_name",
        "项目机构名称": "org_name",
        "上级PJ编码": "parent_pj_code",
        "所在省": "province_code",
        "所在市": "city_code",
        "业务板块": "business_unit",
        "项目承担部门": "dept",
        "项目类型": "project_type",
        "项目组织模式": "org_mode",
        "主数据系统数据状态": "data_status",
        "是否为执行层": "is_execution_level",
        "备注": "remark",
    }

    field_idx = {}
    for idx, header in enumerate(headers):
        field = header_map.get(header)
        if field:
            field_idx[field] = idx

    required_fields = [
        "project_code",
        "project_name",
        "org_name",
        "province_code",
        "business_unit",
        "dept",
        "project_type",
        "org_mode",
        "data_status",
        "is_execution_level",
    ]

    success = 0
    failure = 0

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        for field, idx in field_idx.items():
            row_data[field] = (row[idx] if idx < len(row) else "") or ""

        row_data["project_code"] = _normalize_project_code(row_data.get("project_code"))

        missing = [field for field in required_fields if not row_data.get(field)]
        if missing:
            failure += 1
            continue

        if not PJ_CODE_PATTERN.fullmatch(row_data["project_code"]):
            failure += 1
            continue

        row_data["city_code"] = row_data.get("city_code") or row_data.get("province_code")
        row_data["parent_pj_code"] = row_data.get("parent_pj_code") or None
        row_data["is_execution_level"] = str(row_data.get("is_execution_level")).strip() in [
            "true", "True", "是", "1",
        ]
        if row_data.get("project_code") and len(str(row_data.get("project_code"))) >= 6:
            row_data["project_year"] = str(row_data.get("project_code"))[2:6]
        else:
            row_data["project_year"] = ""
        row_data["created_by"] = submitter
        row_data["updated_by"] = submitter

        try:
            with transaction.atomic():
                # 检查项目是否已存在
                existing = ProjectMaster.objects.filter(project_code=row_data["project_code"]).first()
                if existing:
                    # 更新现有项目
                    for key, value in row_data.items():
                        setattr(existing, key, value)
                    existing.save()
                else:
                    # 创建新项目
                    project = ProjectMaster(**row_data)
                    project.full_clean()
                    project.save()
                
                ProjectMasterLog.objects.create(
                    project_code=row_data["project_code"],
                    action="import",
                    after_data=row_data,
                    operator=submitter,
                    source="approval_import",
                )
            success += 1
        except Exception:
            failure += 1
    
    # 删除临时文件
    os.unlink(file_path)
    
    if failure > 0:
        raise Exception(f"导入完成，成功 {success} 条，失败 {failure} 条")
    
    return success


@login_required
def user_list(request):
    permissions = _get_user_permissions(request.user)
    if not permissions["can_user_manage"]:
        return _redirect_no_permission(request)

    if request.user.is_staff:
        users = User.objects.all().order_by("-date_joined")
    else:
        users = User.objects.filter(id=request.user.id)

    for user in users:
        UserProfile.objects.get_or_create(user=user)

    if request.method == "POST" and request.user.is_staff:
        action = request.POST.get("action")
        if action == "create":
            username = request.POST.get("username", "").strip()
            email = request.POST.get("email", "").strip()
            password = request.POST.get("password", "").strip()
            department = request.POST.get("department", "").strip()
            is_staff = request.POST.get("is_staff") == "true"
            if username and password:
                new_user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    is_staff=is_staff,
                )
                profile, _ = UserProfile.objects.get_or_create(user=new_user)
                profile.department = department
                profile.save(update_fields=["department"])
                messages.success(request, "用户已创建")
            else:
                messages.error(request, "用户名和密码不能为空")
        elif action == "toggle":
            user_id = request.POST.get("user_id")
            target = get_object_or_404(User, id=user_id)
            target.is_active = not target.is_active
            target.save(update_fields=["is_active"])
            messages.success(request, "用户状态已更新")
        elif action == "reset":
            user_id = request.POST.get("user_id")
            new_password = request.POST.get("new_password", "").strip()
            target = get_object_or_404(User, id=user_id)
            if new_password:
                # 密码复杂度验证
                import re
                if len(new_password) < 8:
                    messages.error(request, "密码长度至少8位")
                elif not re.search(r'[A-Z]', new_password):
                    messages.error(request, "密码必须包含大写字母")
                elif not re.search(r'[a-z]', new_password):
                    messages.error(request, "密码必须包含小写字母")
                elif not re.search(r'[0-9]', new_password):
                    messages.error(request, "密码必须包含数字")
                else:
                    target.set_password(new_password)
                    target.save(update_fields=["password"])
                    messages.success(request, f"用户 {target.username} 的密码已重置")
            else:
                messages.error(request, "请输入新密码")
        elif action == "update_department":
            user_id = request.POST.get("user_id")
            department = request.POST.get("department", "").strip()
            target = get_object_or_404(User, id=user_id)
            profile, _ = UserProfile.objects.get_or_create(user=target)
            profile.department = department
            profile.save(update_fields=["department"])
            messages.success(request, f"用户 {target.username} 的部门已更新")
        return redirect("user_list")

    return render(request, "user_list.html", {"users": users, "permissions": permissions})


@login_required
def permission_manage(request):
    permissions = _get_user_permissions(request.user)
    if not request.user.is_staff or not permissions["can_user_manage"]:
        return _redirect_no_permission(request)

    if request.method == "POST":
        user_id = request.POST.get("user_id")
        target = get_object_or_404(User, id=user_id)
        profile, _ = UserProfile.objects.get_or_create(user=target)
        for key, _ in PERMISSION_FIELDS:
            setattr(profile, key, request.POST.get(key) == "on")
        profile.save(update_fields=[key for key, _ in PERMISSION_FIELDS])
        messages.success(request, f"用户 {target.username} 的权限已更新")
        return redirect("permission_manage")

    users = list(User.objects.all().order_by("-date_joined"))
    user_permission_rows = []
    for user in users:
        profile, _ = UserProfile.objects.get_or_create(user=user)
        user_permission_rows.append(
            {
                "user": user,
                "profile": profile,
            }
        )

    return render(
        request,
        "permission_manage.html",
        {
            "user_permission_rows": user_permission_rows,
            "permissions": permissions,
        },
    )

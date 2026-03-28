import os
import re
import tempfile
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import transaction
from django.db.models import Q, Sum, Count, CharField
from django.db.models.functions import Cast
from django.forms.models import model_to_dict
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    ADJUSTMENT_TYPE_CHOICES,
    APPROVAL_STATUS_CHOICES,
    CONTRACT_CATEGORY_CHOICES,
    CONTRACT_DIRECTION_CHOICES,
    CONTRACT_STATUS_CHOICES,
    DictType,
    PARTY_TYPE_CHOICES,
    SOURCE_SYSTEM_CHOICES,
    ContractAdjustment,
    ContractAdjustmentActionLog,
    ContractMaster,
    Counterparty,
    ProjectApproval,
    ProjectMaster,
)

CT_CODE_PATTERN = re.compile(r"^CT\d{12}$")


def _get_permissions(user):
    from .views import _get_user_permissions
    return _get_user_permissions(user)


def _get_dept_name_map():
    dept_type = DictType.objects.filter(code="DEPT", is_active=True).prefetch_related("items").first()
    if not dept_type:
        return {}
    return {
        item.code: item.name
        for item in dept_type.items.all()
        if item.is_active
    }


def _to_decimal(value, default="0"):
    text = str(value if value is not None else "").strip()
    if text == "":
        text = default
    return Decimal(text)


def _serialize_counterparty(counterparty):
    data = model_to_dict(
        counterparty,
        fields=[
            "party_name",
            "party_type",
            "credit_code",
            "contact_name",
            "contact_phone",
            "status",
            "remark",
            "established_date",
            "province_code",
            "city",
            "enterprise_type",
            "industry",
            "former_name",
            "registration_address",
            "business_scope",
        ],
    )
    if data.get("established_date"):
        data["established_date"] = data["established_date"].strftime("%Y-%m-%d")
    return data


def _serialize_contract(contract):
    data = model_to_dict(
        contract,
        fields=[
            "project",
            "execution_project",
            "counterparty",
            "contract_name",
            "contract_no",
            "source_system",
            "contract_direction",
            "contract_category",
            "contract_year",
            "contract_status",
            "remark",
            "sign_date",
            "effective_date",
            "close_date",
        ],
    )
    for date_key in ["sign_date", "effective_date", "close_date"]:
        if data.get(date_key):
            data[date_key] = data[date_key].strftime("%Y-%m-%d")
    data["project_id"] = data.pop("project", None)
    data["execution_project_id"] = data.pop("execution_project", None)
    data["counterparty_id"] = data.pop("counterparty", None)
    return data


def _get_counterparty_province_data():
    province_dict = DictType.objects.filter(code="PROVINCE", is_active=True).prefetch_related("items").first()
    province_items = list(province_dict.items.filter(is_active=True).order_by("sort_order", "code")) if province_dict else []
    province_name_map = {item.code: item.name for item in province_items}
    return province_items, province_name_map


def _apply_counterparty_filters(queryset, filters, province_items):
    keyword = filters["keyword"]
    if keyword:
        queryset = queryset.filter(
            Q(party_name__icontains=keyword)
            | Q(contact_name__icontains=keyword)
            | Q(contact_phone__icontains=keyword)
            | Q(registration_address__icontains=keyword)
            | Q(former_name__icontains=keyword)
            | Q(business_scope__icontains=keyword)
        )

    if filters["province"]:
        queryset = queryset.filter(province_code=filters["province"])

    if filters.get("party_name"):
        queryset = queryset.filter(party_name__icontains=filters["party_name"])

    return queryset


def _decorate_counterparties(counterparties, province_name_map):
    party_type_lookup = {value: label for value, label in PARTY_TYPE_CHOICES}
    for item in counterparties:
        item.province_name = province_name_map.get(item.province_code, item.province_code or "—")
        item.party_type_name = party_type_lookup.get(item.party_type, item.party_type)
        item.business_scope_short = (item.business_scope or "")[:20]
        item.business_scope_truncated = bool(item.business_scope and len(item.business_scope) > 20)
    return counterparties


def _has_pending_approval(target_module, target_id, approval_type):
    pending = ProjectApproval.objects.filter(approval_type=approval_type, status="pending")
    for item in pending:
        after_data = item.after_data or {}
        if after_data.get("target_module") == target_module and str(after_data.get("target_id")) == str(target_id):
            return item
    return None


def _submit_business_approval(request, approval_type, target_module, target_id, target_code, target_name, before_data=None, after_data=None, change_note=""):
    # ProjectApproval.project_code/project_name are legacy fields with fixed lengths.
    # Normalize values to avoid DB write errors when approving non-project modules.
    safe_name = str(target_name or "")[:200]

    payload = {
        "target_module": target_module,
        "target_id": target_id,
        "target_code": target_code,
        "target_name": target_name,
    }
    if after_data:
        payload.update(after_data)

    return ProjectApproval.objects.create(
        project_code="",
        project_name=safe_name,
        approval_type=approval_type,
        before_data=before_data,
        after_data=payload,
        change_note=change_note,
        submitter=request.user.username,
        approver="倪明珠",
        status="pending",
    )


@login_required
def export_counterparty_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "往来单位导入模板"
    ws.append([
        "单位名称",
        "单位类型",
        "统一社会信用代码",
        "联系人",
        "联系电话",
        "状态",
        "备注",
        "成立日期",
        "所属省份",
        "所属城市",
        "企业类型",
        "所属行业",
        "曾用名",
        "注册地址",
        "经营范围",
    ])

    ws_ref = wb.create_sheet(title="字典参考")
    ws_ref.append(["字段", "可选值", "说明"])
    ws_ref.append(["单位类型", "OWNER", "业主"])
    ws_ref.append(["单位类型", "SUPPLIER", "供应商"])
    ws_ref.append(["单位类型", "SUBCONTRACTOR", "分包商"])
    ws_ref.append(["单位类型", "SUPPLY_SUB", "供应&分包商"])
    ws_ref.append(["单位类型", "OTHER_VENDOR", "其他外委单位"])
    ws_ref.append(["状态", "ACTIVE", "启用"])
    ws_ref.append(["状态", "INACTIVE", "停用"])
    
    # 前段省份列表
    province_dict = DictType.objects.filter(code="PROVINCE", is_active=True).first()
    if province_dict:
        for item in province_dict.items.filter(is_active=True).order_by("sort_order", "code")[:50]:
            ws_ref.append(["所属省份", item.code, item.name])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 30
    ws.column_dimensions["O"].width = 30
    ws_ref.column_dimensions["A"].width = 16
    ws_ref.column_dimensions["B"].width = 20
    ws_ref.column_dimensions["C"].width = 28

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=counterparty_import_template.xlsx"
    wb.save(response)
    return response


@login_required
def export_counterparty_list(request):
    permissions = _get_permissions(request.user)
    if not permissions.get("can_manage_counterparty"):
        from .views import _redirect_no_permission
        return _redirect_no_permission(request)

    from .views import _ensure_export_approved

    if _ensure_export_approved(request, "counterparty_list", "往来单位列表导出") is None:
        return redirect("contract_counterparty_list")

    filters = {
        "keyword": request.GET.get("keyword", "").strip(),
        "province": request.GET.get("province", "").strip(),
        "city": request.GET.get("city", "").strip(),
    }
    province_items, province_name_map = _get_counterparty_province_data()
    qs = _apply_counterparty_filters(Counterparty.objects.all(), filters, province_items).order_by("party_name")
    counterparties = _decorate_counterparties(list(qs), province_name_map)

    wb = Workbook()
    ws = wb.active
    ws.title = "往来单位列表"
    headers = [
        "单位名称",
        "单位类型",
        "统一社会信用代码",
        "联系人",
        "联系电话",
        "状态",
        "成立日期",
        "所属省份",
        "所属城市",
        "企业类型",
        "所属行业",
        "曾用名",
        "注册地址",
        "经营范围",
        "备注",
        "创建时间",
        "创建人",
    ]
    ws.append(headers)
    for item in counterparties:
        ws.append([
            item.party_name,
            item.party_type_name,
            item.credit_code,
            item.contact_name,
            item.contact_phone,
            "启用" if item.status == "ACTIVE" else "停用",
            item.established_date.strftime("%Y-%m-%d") if item.established_date else "",
            item.province_name if item.province_name != "—" else "",
            item.city,
            item.enterprise_type,
            item.industry,
            item.former_name,
            item.registration_address,
            item.business_scope,
            item.remark,
            item.created_at.strftime("%Y-%m-%d") if item.created_at else "",
            item.created_by,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=counterparty_list.xlsx"
    wb.save(response)
    return response


@login_required
def export_contract_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "合同台账导入模板"
    ws.append([
        "项目主数据编码",
        "合同CT码",
        "合同名称",
        "统一社会信用代码",
        "合同编号",
        "来源系统",
        "合同方向",
        "合同分类",
        "合同年份",
        "原始含税金额",
        "原始不含税金额",
        "原始税率",
        "当前含税金额",
        "当前不含税金额",
        "当前税率",
        "合同状态",
        "备注",
    ])

    ws_ref = wb.create_sheet(title="字典参考")
    ws_ref.append(["字段", "可选值", "说明"])
    for value, label in SOURCE_SYSTEM_CHOICES:
        ws_ref.append(["来源系统", value, label])
    for value, label in CONTRACT_DIRECTION_CHOICES:
        ws_ref.append(["合同方向", value, label])
    for value, label in CONTRACT_CATEGORY_CHOICES:
        ws_ref.append(["合同分类", value, label])
    for value, label in CONTRACT_STATUS_CHOICES:
        ws_ref.append(["合同状态", value, label])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 12
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 30
    ws_ref.column_dimensions["A"].width = 16
    ws_ref.column_dimensions["B"].width = 20
    ws_ref.column_dimensions["C"].width = 28

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=contract_import_template.xlsx"
    wb.save(response)
    return response


def _log_adjustment_action(adjustment, action_type, user, comment="", from_status="", to_status=""):
    ContractAdjustmentActionLog.objects.create(
        adjustment=adjustment,
        action_type=action_type,
        action_by=user,
        comment=comment,
        from_status=from_status,
        to_status=to_status,
    )


def _apply_adjustment_to_contract(adjustment):
    contract = adjustment.contract
    contract.current_amount_tax = adjustment.after_amount_tax
    contract.current_amount_notax = adjustment.after_amount_notax
    contract.current_tax_rate = adjustment.after_tax_rate
    if adjustment.after_counterparty_id:
        contract.counterparty = adjustment.after_counterparty
        contract.counterparty_name_snapshot = adjustment.after_counterparty_name
    contract.approved_adjustment_count = contract.approved_adjustment_count + 1
    contract.last_adjustment_date = adjustment.adjustment_date
    contract.save()


# ---------------------------------------------------------------------------
# 往来单位管理
# ---------------------------------------------------------------------------

@login_required
def contract_counterparty_view(request):
    permissions = _get_permissions(request.user)
    if not permissions.get("can_manage_counterparty"):
        from .views import _redirect_no_permission
        return _redirect_no_permission(request)

    province_items, province_name_map = _get_counterparty_province_data()

    if request.method == "POST":
        form_type = request.POST.get("form_type", "").strip()

        if form_type == "create_counterparty":
            credit_code = request.POST.get("credit_code", "").strip().upper()
            contact_phone = request.POST.get("contact_phone", "").strip()[:255]
            if len(credit_code) != 18:
                messages.error(request, "统一社会信用代码必须为18位")
                return redirect("contract_counterparty_list")
            if Counterparty.objects.filter(credit_code=credit_code).exists():
                messages.error(request, "统一社会信用代码已存在，请确保唯一")
                return redirect("contract_counterparty_list")

            established_date_str = request.POST.get("established_date", "").strip()
            established_date = None
            if established_date_str:
                try:
                    established_date = datetime.strptime(established_date_str, "%Y-%m-%d").date()
                except ValueError:
                    pass

            try:
                with transaction.atomic():
                    Counterparty.objects.create(
                        party_name=request.POST.get("party_name", "").strip(),
                        party_type=request.POST.get("party_type", "").strip(),
                        credit_code=credit_code,
                        contact_name=request.POST.get("contact_name", "").strip(),
                        contact_phone=contact_phone,
                        status=request.POST.get("status", "ACTIVE").strip() or "ACTIVE",
                        remark=request.POST.get("remark", "").strip(),
                        established_date=established_date,
                        province_code=request.POST.get("province_code", "").strip(),
                        city=request.POST.get("city", "").strip(),
                        enterprise_type=request.POST.get("enterprise_type", "").strip(),
                        industry=request.POST.get("industry", "").strip(),
                        former_name=request.POST.get("former_name", "").strip(),
                        registration_address=request.POST.get("registration_address", "").strip(),
                        business_scope=request.POST.get("business_scope", "").strip(),
                        created_by=request.user.username,
                        updated_by=request.user.username,
                    )
                messages.success(request, "往来单位已新增")
            except Exception as exc:
                messages.error(request, f"往来单位保存失败：{exc}")
            return redirect("contract_counterparty_list")

        if form_type == "update_counterparty":
            counterparty_id = request.POST.get("counterparty_id", "").strip()
            target = get_object_or_404(Counterparty, id=counterparty_id)
            existing = _has_pending_approval("counterparty", target.id, "update")
            if existing:
                messages.warning(request, f"该往来单位已有待审批修改申请，审批单号：{existing.id}")
                return redirect("contract_counterparty_list")

            contact_phone = request.POST.get("contact_phone", target.contact_phone).strip()[:255]

            after_data = {
                "party_name": request.POST.get("party_name", target.party_name).strip(),
                "party_type": request.POST.get("party_type", target.party_type).strip(),
                "contact_name": request.POST.get("contact_name", target.contact_name).strip(),
                "contact_phone": contact_phone,
                "status": request.POST.get("status", target.status).strip(),
                "remark": request.POST.get("remark", target.remark).strip(),
            }
            approval = _submit_business_approval(
                request,
                approval_type="update",
                target_module="counterparty",
                target_id=target.id,
                target_code=target.credit_code,
                target_name=target.party_name,
                before_data=_serialize_counterparty(target),
                after_data=after_data,
                change_note=request.POST.get("change_note", "").strip() or "往来单位修改申请",
            )
            messages.success(request, f"往来单位修改申请已提交，等待倪明珠审批。审批单号：{approval.id}")
            return redirect("contract_counterparty_list")

        if form_type == "delete_counterparty":
            counterparty_id = request.POST.get("counterparty_id", "").strip()
            target = get_object_or_404(Counterparty, id=counterparty_id)
            existing = _has_pending_approval("counterparty", target.id, "delete")
            if existing:
                messages.warning(request, f"该往来单位已有待审批删除申请，审批单号：{existing.id}")
                return redirect("contract_counterparty_list")

            approval = _submit_business_approval(
                request,
                approval_type="delete",
                target_module="counterparty",
                target_id=target.id,
                target_code=target.credit_code,
                target_name=target.party_name,
                before_data=_serialize_counterparty(target),
                change_note=request.POST.get("change_note", "").strip() or "往来单位删除申请",
            )
            messages.success(request, f"往来单位删除申请已提交，等待倪明珠审批。审批单号：{approval.id}")
            return redirect("contract_counterparty_list")

    filters = {
        "keyword": request.GET.get("keyword", "").strip(),
        "province": request.GET.get("province", "").strip(),
        "party_name": request.GET.get("party_name", "").strip(),
    }
    qs = _apply_counterparty_filters(Counterparty.objects.all(), filters, province_items).order_by("-updated_at", "-id")
    
    # 标准化分页处理（完整的异常保护）
    paginator = Paginator(qs, 50)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    counterparties = _decorate_counterparties(list(page_obj.object_list), province_name_map)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    export_query_string = urlencode({
        "keyword": filters["keyword"],
        "province": filters["province"],
        "party_name": filters["party_name"],
    })

    context = {
        "counterparties": counterparties,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": page_obj.has_other_pages(),
        "current_query": query_params.urlencode(),
        "export_query_string": export_query_string,
        "filters": filters,
        "permissions": permissions,
        "party_type_choices": PARTY_TYPE_CHOICES,
        "province_items": province_items,
        "active_menu": "contract_counterparty",
        "total_count": qs.count(),
    }
    return render(request, "contract_counterparty.html", context)


@login_required
def contract_counterparty_edit(request, counterparty_id):
    permissions = _get_permissions(request.user)
    if not permissions.get("can_manage_counterparty"):
        from .views import _redirect_no_permission
        return _redirect_no_permission(request)

    counterparty = get_object_or_404(Counterparty, id=counterparty_id)
    province_items, province_name_map = _get_counterparty_province_data()
    
    # 添加province_name到counterparty
    counterparty.province_name = province_name_map.get(counterparty.province_code, counterparty.province_code or "")
    
    if request.method == "POST":
        form_type = request.POST.get("form_type", "").strip()
        
        if form_type == "update_counterparty":
            # 检查是否有待审批的修改
            existing = _has_pending_approval("counterparty", counterparty.id, "update")
            if existing:
                messages.warning(request, f"该往来单位已有待审批修改申请，审批单号：{existing.id}")
                return redirect("contract_counterparty_edit", counterparty_id=counterparty_id)
            
            contact_phone = request.POST.get("contact_phone", counterparty.contact_phone).strip()[:255]
            
            after_data = {
                "party_name": request.POST.get("party_name", counterparty.party_name).strip(),
                "party_type": request.POST.get("party_type", counterparty.party_type).strip(),
                "contact_name": request.POST.get("contact_name", counterparty.contact_name).strip(),
                "contact_phone": contact_phone,
                "status": request.POST.get("status", counterparty.status).strip(),
                "remark": request.POST.get("remark", counterparty.remark).strip(),
                "established_date": request.POST.get("established_date", "").strip(),
                "province_code": request.POST.get("province_code", counterparty.province_code).strip(),
                "city": request.POST.get("city", counterparty.city).strip(),
                "enterprise_type": request.POST.get("enterprise_type", counterparty.enterprise_type).strip(),
                "industry": request.POST.get("industry", counterparty.industry).strip(),
                "former_name": request.POST.get("former_name", counterparty.former_name).strip(),
                "registration_address": request.POST.get("registration_address", counterparty.registration_address).strip(),
                "business_scope": request.POST.get("business_scope", counterparty.business_scope).strip(),
            }
            
            approval = _submit_business_approval(
                request,
                approval_type="update",
                target_module="counterparty",
                target_id=counterparty.id,
                target_code=counterparty.credit_code,
                target_name=counterparty.party_name,
                before_data=_serialize_counterparty(counterparty),
                after_data=after_data,
                change_note=request.POST.get("change_note", "").strip() or "往来单位修改申请",
            )
            messages.success(request, f"往来单位修改申请已提交，等待审批。审批单号：{approval.id}")
            return redirect("contract_counterparty_list")
    
    # 构建party_type_display
    party_type_lookup = {value: label for value, label in PARTY_TYPE_CHOICES}
    counterparty.party_type_display = party_type_lookup.get(counterparty.party_type, counterparty.party_type)
    
    context = {
        "counterparty": counterparty,
        "province_items": province_items,
        "party_type_choices": PARTY_TYPE_CHOICES,
    }
    return render(request, "contract_counterparty_edit.html", context)


# ---------------------------------------------------------------------------
# 合同台账管理
# ---------------------------------------------------------------------------

@login_required
def contract_list_view(request):
    permissions = _get_permissions(request.user)
    if not permissions.get("can_view_contract_ledger"):
        from .views import _redirect_no_permission
        return _redirect_no_permission(request)
    dept_name_map = _get_dept_name_map()
    execution_projects = list(
        ProjectMaster.objects.filter(is_deleted=False, is_execution_level=True).order_by("-project_code")
    )
    for ep in execution_projects:
        ep.dept_name = dept_name_map.get(ep.dept, ep.dept or "")

    if request.method == "POST":
        form_type = request.POST.get("form_type", "").strip()

        if form_type == "create_contract":
            project_id = request.POST.get("project_id", "").strip()
            execution_project_id = request.POST.get("execution_project_id", "").strip()
            counterparty_id = request.POST.get("counterparty_id", "").strip()
            ct_code = request.POST.get("contract_ct_code", "").strip().upper()
            if not CT_CODE_PATTERN.fullmatch(ct_code):
                messages.error(request, "CT码格式错误，必须为CT+12位数字")
                return redirect("contract_list")
            if not execution_project_id:
                messages.error(request, "对应执行层项目为必填项")
                return redirect("contract_list")
            try:
                project = ProjectMaster.objects.get(id=project_id, is_deleted=False)
                execution_project = ProjectMaster.objects.get(id=execution_project_id, is_deleted=False, is_execution_level=True)
                counterparty = Counterparty.objects.get(id=counterparty_id)
                with transaction.atomic():
                    contract = ContractMaster(
                        project=project,
                        execution_project=execution_project,
                        counterparty=counterparty,
                        contract_ct_code=ct_code,
                        contract_name=request.POST.get("contract_name", "").strip(),
                        contract_no=request.POST.get("contract_no", "").strip(),
                        source_system=request.POST.get("source_system", "").strip(),
                        source_record_id=request.POST.get("source_record_id", "").strip(),
                        source_contract_no=request.POST.get("source_contract_no", "").strip(),
                        contract_direction=request.POST.get("contract_direction", "").strip(),
                        contract_category=request.POST.get("contract_category", "").strip(),
                        undertaking_dept_code="",
                        undertaking_dept_name=dept_name_map.get(execution_project.dept, execution_project.dept or ""),
                        contract_year=request.POST.get("contract_year", "").strip(),
                        sign_date=request.POST.get("sign_date") or None,
                        effective_date=request.POST.get("effective_date") or None,
                        close_date=request.POST.get("close_date") or None,
                        original_amount_tax=_to_decimal(request.POST.get("original_amount_tax", "0")),
                        original_amount_notax=_to_decimal(request.POST.get("original_amount_notax", "0")),
                        original_tax_rate=_to_decimal(request.POST.get("original_tax_rate"), default="0") if request.POST.get("original_tax_rate", "").strip() else None,
                        current_amount_tax=_to_decimal(request.POST.get("current_amount_tax", "0")),
                        current_amount_notax=_to_decimal(request.POST.get("current_amount_notax", "0")),
                        current_tax_rate=_to_decimal(request.POST.get("current_tax_rate"), default="0") if request.POST.get("current_tax_rate", "").strip() else None,
                        contract_status=request.POST.get("contract_status", "SIGNED").strip() or "SIGNED",
                        remark=request.POST.get("remark", "").strip(),
                        created_by=request.user.username,
                        updated_by=request.user.username,
                    )
                    if contract.current_amount_tax == Decimal("0") and contract.current_amount_notax == Decimal("0"):
                        contract.current_amount_tax = contract.original_amount_tax
                        contract.current_amount_notax = contract.original_amount_notax
                        if not contract.current_tax_rate:
                            contract.current_tax_rate = contract.original_tax_rate
                    contract.full_clean()
                    contract.save()
                messages.success(request, "合同已新增")
            except Exception as exc:
                messages.error(request, f"合同保存失败：{exc}")
            return redirect("contract_list")

        if form_type == "update_contract":
            contract_id = request.POST.get("contract_id", "").strip()
            target = get_object_or_404(ContractMaster, id=contract_id, is_deleted=False)
            existing = _has_pending_approval("contract", target.id, "update")
            if existing:
                messages.warning(request, f"该合同已有待审批修改申请，审批单号：{existing.id}")
                return redirect("contract_list")

            after_data = {
                "contract_name": request.POST.get("contract_name", target.contract_name).strip(),
                "contract_status": request.POST.get("contract_status", target.contract_status).strip(),
                "remark": request.POST.get("remark", target.remark).strip(),
            }
            approval = _submit_business_approval(
                request,
                approval_type="update",
                target_module="contract",
                target_id=target.id,
                target_code=target.contract_ct_code,
                target_name=target.contract_name,
                before_data=_serialize_contract(target),
                after_data=after_data,
                change_note=request.POST.get("change_note", "").strip() or "合同修改申请",
            )
            messages.success(request, f"合同修改申请已提交，等待倪明珠审批。审批单号：{approval.id}")
            return redirect("contract_list")

        if form_type == "delete_contract":
            contract_id = request.POST.get("contract_id", "").strip()
            target = get_object_or_404(ContractMaster, id=contract_id, is_deleted=False)
            existing = _has_pending_approval("contract", target.id, "delete")
            if existing:
                messages.warning(request, f"该合同已有待审批删除申请，审批单号：{existing.id}")
                return redirect("contract_list")

            approval = _submit_business_approval(
                request,
                approval_type="delete",
                target_module="contract",
                target_id=target.id,
                target_code=target.contract_ct_code,
                target_name=target.contract_name,
                before_data=_serialize_contract(target),
                change_note=request.POST.get("change_note", "").strip() or "合同删除申请",
            )
            messages.success(request, f"合同删除申请已提交，等待倪明珠审批。审批单号：{approval.id}")
            return redirect("contract_list")

    filters = {
        "project_code": request.GET.get("project_code", "").strip(),
        "contract_ct_code": request.GET.get("contract_ct_code", "").strip(),
        "contract_name": request.GET.get("contract_name", "").strip(),
        "source_system": request.GET.get("source_system", "").strip(),
        "contract_direction": request.GET.get("contract_direction", "").strip(),
        "contract_category": request.GET.get("contract_category", "").strip(),
        "contract_status": request.GET.get("contract_status", "").strip(),
        "undertaking_dept": request.GET.get("undertaking_dept", "").strip(),
        "contract_year": request.GET.get("contract_year", "").strip(),
        "counterparty_name": request.GET.get("counterparty_name", "").strip(),
    }
    qs = ContractMaster.objects.filter(is_deleted=False).select_related("project", "counterparty")
    if filters["project_code"]:
        qs = qs.filter(project_code_snapshot__icontains=filters["project_code"])
    if filters["contract_ct_code"]:
        qs = qs.filter(contract_ct_code__icontains=filters["contract_ct_code"])
    if filters["contract_name"]:
        qs = qs.filter(contract_name__icontains=filters["contract_name"])
    if filters["source_system"]:
        qs = qs.filter(source_system=filters["source_system"])
    if filters["contract_direction"]:
        qs = qs.filter(contract_direction=filters["contract_direction"])
    if filters["contract_category"]:
        qs = qs.filter(contract_category=filters["contract_category"])
    if filters["contract_status"]:
        qs = qs.filter(contract_status=filters["contract_status"])
    if filters["undertaking_dept"]:
        qs = qs.filter(undertaking_dept_name__icontains=filters["undertaking_dept"])
    if filters["contract_year"]:
        qs = qs.filter(contract_year__icontains=filters["contract_year"])
    if filters["counterparty_name"]:
        qs = qs.filter(counterparty_name_snapshot__icontains=filters["counterparty_name"])

    projects = list(ProjectMaster.objects.filter(is_deleted=False).order_by("-project_code"))
    for project in projects:
        project.dept_name = dept_name_map.get(project.dept, project.dept or "")
    
    # 执行分页（稳定排序 + 完整的页码保护）
    ordered_qs = qs.order_by("-created_at", "-id")  # 二级排序确保稳定性
    paginator = Paginator(ordered_qs, 50)  # 每页50条
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        "contracts": list(page_obj.object_list),
        "page_obj": page_obj,
        "paginator": paginator,
        "projects": projects,
        "execution_projects": execution_projects,
        "counterparties": Counterparty.objects.filter(status="ACTIVE").order_by("party_name"),
        "filters": filters,
        "permissions": permissions,
        "source_system_choices": SOURCE_SYSTEM_CHOICES,
        "contract_direction_choices": CONTRACT_DIRECTION_CHOICES,
        "contract_category_choices": CONTRACT_CATEGORY_CHOICES,
        "contract_status_choices": CONTRACT_STATUS_CHOICES,
        "active_menu": "contract_list",
        "total_count": qs.count(),
    }
    return render(request, "contract_list.html", context)


# ---------------------------------------------------------------------------
# 合同调整管理
# ---------------------------------------------------------------------------

@login_required
def contract_adjustment_view(request):
    permissions = _get_permissions(request.user)
    if not permissions.get("can_edit_contract_adjustment"):
        from .views import _redirect_no_permission
        return _redirect_no_permission(request)

    if request.method == "POST":
        form_type = request.POST.get("form_type", "").strip()

        if form_type == "create_adjustment":
            contract_id = request.POST.get("contract_id", "").strip()
            try:
                contract = ContractMaster.objects.get(id=contract_id, is_deleted=False)
                with transaction.atomic():
                    adjustment = ContractAdjustment(
                        contract=contract,
                        project=contract.project,
                        adjustment_type=request.POST.get("adjustment_type", "").strip(),
                        adjustment_no=request.POST.get("adjustment_no", "").strip(),
                        adjustment_date=request.POST.get("adjustment_date"),
                        effective_date=request.POST.get("effective_date") or None,
                        change_amount_tax=_to_decimal(request.POST.get("change_amount_tax", "0")),
                        change_amount_notax=_to_decimal(request.POST.get("change_amount_notax", "0")),
                        after_tax_rate=_to_decimal(request.POST.get("after_tax_rate"), default="0") if request.POST.get("after_tax_rate", "").strip() else None,
                        after_counterparty_id=request.POST.get("after_counterparty_id") or None,
                        remark=request.POST.get("remark", "").strip(),
                        approval_status=request.POST.get("approval_status", "DRAFT").strip() or "DRAFT",
                        submitted_by=request.user,
                        submitted_at=timezone.now(),
                        approver_name="\u502a\u660e\u73e0",
                        source_system=request.POST.get("source_system", contract.source_system).strip() or contract.source_system,
                        source_record_id=request.POST.get("source_record_id", "").strip(),
                        created_by=request.user.username,
                        updated_by=request.user.username,
                    )
                    adjustment.save()
                    _log_adjustment_action(
                        adjustment, "EDIT", request.user,
                        comment="\u521b\u5efa\u8c03\u6574\u8bb0\u5f55", from_status="", to_status=adjustment.approval_status,
                    )
                    if adjustment.approval_status == "APPROVED":
                        _apply_adjustment_to_contract(adjustment)
                        _log_adjustment_action(
                            adjustment, "APPROVE", request.user,
                            comment="\u5f55\u5165\u5373\u901a\u8fc7\u5e76\u540c\u6b65\u4e3b\u8868",
                            from_status="APPROVED", to_status="APPROVED",
                        )
                messages.success(request, "\u8c03\u6574\u8bb0\u5f55\u5df2\u4fdd\u5b58")
            except Exception as exc:
                messages.error(request, f"\u8c03\u6574\u8bb0\u5f55\u4fdd\u5b58\u5931\u8d25\uff1a{exc}")
            return redirect("contract_adjustment_list")

        if form_type == "update_adjustment_status":
            adjustment_id = request.POST.get("adjustment_id", "").strip()
            to_status = request.POST.get("to_status", "").strip()
            comment = request.POST.get("comment", "").strip()
            return_reason = request.POST.get("return_reason", "").strip()
            try:
                adjustment = ContractAdjustment.objects.select_related("contract").get(id=adjustment_id)
                from_status = adjustment.approval_status
                with transaction.atomic():
                    adjustment.approval_status = to_status
                    adjustment.approval_comment = comment
                    if to_status == "RETURNED":
                        adjustment.return_reason = return_reason
                    if to_status == "IN_REVIEW":
                        adjustment.submitted_by = request.user
                        adjustment.submitted_at = timezone.now()
                    if to_status == "APPROVED":
                        adjustment.approver = request.user
                        adjustment.approved_at = timezone.now()
                    adjustment.updated_by = request.user.username
                    adjustment.save()
                    if to_status == "APPROVED" and from_status != "APPROVED":
                        _apply_adjustment_to_contract(adjustment)
                    action_type = "EDIT"
                    if to_status == "IN_REVIEW":
                        action_type = "SUBMIT"
                    elif to_status == "APPROVED":
                        action_type = "APPROVE"
                    elif to_status == "RETURNED":
                        action_type = "RETURN"
                    _log_adjustment_action(
                        adjustment, action_type, request.user,
                        comment=comment, from_status=from_status, to_status=to_status,
                    )
                messages.success(request, "\u8c03\u6574\u72b6\u6001\u5df2\u66f4\u65b0")
            except Exception as exc:
                messages.error(request, f"\u8c03\u6574\u72b6\u6001\u66f4\u65b0\u5931\u8d25\uff1a{exc}")
            return redirect("contract_adjustment_list")

    filters = {
        "adj_ct_code": request.GET.get("adj_ct_code", "").strip(),
        "adjustment_type": request.GET.get("adjustment_type", "").strip(),
        "approval_status": request.GET.get("approval_status", "").strip(),
        "adjustment_no": request.GET.get("adjustment_no", "").strip(),
    }
    qs = ContractAdjustment.objects.select_related(
        "contract", "project", "before_counterparty", "after_counterparty"
    )
    if filters["adj_ct_code"]:
        qs = qs.filter(contract_ct_code_snapshot__icontains=filters["adj_ct_code"])
    if filters["adjustment_type"]:
        qs = qs.filter(adjustment_type=filters["adjustment_type"])
    if filters["approval_status"]:
        qs = qs.filter(approval_status=filters["approval_status"])
    if filters["adjustment_no"]:
        qs = qs.filter(adjustment_no__icontains=filters["adjustment_no"])

    # 标准化分页处理（完整的异常保护）
    ordered_qs = qs.order_by("-created_at", "-id")
    paginator = Paginator(ordered_qs, 50)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        "page_obj": page_obj,
        "paginator": paginator,
        "adjustments": list(page_obj.object_list),
        "contracts": ContractMaster.objects.filter(is_deleted=False).order_by("-contract_ct_code"),
        "counterparties": Counterparty.objects.filter(status="ACTIVE").order_by("party_name"),
        "filters": filters,
        "permissions": permissions,
        "adjustment_type_choices": ADJUSTMENT_TYPE_CHOICES,
        "approval_status_choices": APPROVAL_STATUS_CHOICES,
        "source_system_choices": SOURCE_SYSTEM_CHOICES,
        "in_review_count": qs.filter(approval_status="IN_REVIEW").count(),
        "returned_count": qs.filter(approval_status="RETURNED").count(),
        "active_menu": "contract_adjustment",
    }
    return render(request, "contract_adjustment.html", context)


# ---------------------------------------------------------------------------
# 导入视图
# ---------------------------------------------------------------------------

@login_required
def import_counterparty_ledger(request):
    if request.method != "POST":
        return redirect("contract_counterparty_list")

    mode = request.POST.get("mode", "insert").strip().lower()
    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "\u8bf7\u9009\u62e9\u5f80\u6765\u5355\u4f4d\u53f0\u8d26\u6587\u4ef6")
        return redirect("contract_counterparty_list")

    if mode not in {"insert", "upsert"}:
        messages.error(request, "\u5bfc\u5165\u6a21\u5f0f\u4e0d\u5408\u6cd5")
        return redirect("contract_counterparty_list")

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    mapping = {
        "\u5355\u4f4d\u540d\u79f0": "party_name",
        "\u5355\u4f4d\u7c7b\u578b": "party_type",
        "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801": "credit_code",
        "\u8054\u7cfb\u4eba": "contact_name",
        "\u8054\u7cfb\u7535\u8bdd": "contact_phone",
        "\u72b6\u6001": "status",
        "\u5907\u6ce8": "remark",
        "\u6210\u7acb\u65e5\u671f": "established_date",
        "\u6240\u5c5e\u7701\u4efd": "province_code",
        "\u6240\u5c5e\u57ce\u5e02": "city",
        "\u4f01\u4e1a\u7c7b\u578b": "enterprise_type",
        "\u6240\u5c5e\u884c\u4e1a": "industry",
        "\u66fe\u7528\u540d": "former_name",
        "\u6ce8\u518c\u5730\u5740": "registration_address",
        "\u7ecf\u8425\u8303\u56f4": "business_scope",
    }
    idx_map = {}
    for idx, header in enumerate(headers):
        if header in mapping:
            idx_map[mapping[header]] = idx

    for field in ["party_name", "party_type", "credit_code"]:
        if field not in idx_map:
            messages.error(request, f"\u5f80\u6765\u5355\u4f4d\u5bfc\u5165\u7f3a\u5c11\u5fc5\u8981\u5217\uff1a{field}")
            return redirect("contract_counterparty_list")

    invalid_credit_code_rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {
            field: str(row[idx] if idx < len(row) and row[idx] is not None else "").strip()
            for field, idx in idx_map.items()
        }
        code = row_data.get("credit_code", "").upper()
        if code and len(code) != 18:
            invalid_credit_code_rows.append(row_no)

    if invalid_credit_code_rows:
        row_text = "\u3001".join(str(i) for i in invalid_credit_code_rows)
        messages.error(request, f"\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u780118\u4f4d\u957f\u5ea6\u4e0d\u5408\u6cd5\u884c\u53f7\uff1a\u7b2c {row_text} \u884c")
        return redirect("contract_counterparty_list")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        file.seek(0)
        for chunk in file.chunks():
            tmp_file.write(chunk)
        tmp_path = tmp_file.name

    approval = _submit_business_approval(
        request,
        approval_type="import",
        target_module="counterparty",
        target_id=0,
        target_code=f"CP-IM-{uuid.uuid4().hex[:8].upper()}",
        target_name="\u5f80\u6765\u5355\u4f4d\u53f0\u8d26\u5bfc\u5165",
        after_data={"mode": mode, "file_name": file.name},
        change_note=f"\u5f80\u6765\u5355\u4f4d\u5bfc\u5165\u7533\u8bf7\uff08{mode}\uff09",
    )
    approval.import_file_path = tmp_path
    approval.save(update_fields=["import_file_path"])
    messages.success(request, f"\u5bfc\u5165\u5ba1\u6279\u5df2\u63d0\u4ea4\uff0c\u7b49\u5f85\u502a\u660e\u73e0\u5ba1\u6279\u3002\u5ba1\u6279\u5355\u53f7\uff1a{approval.id}")
    return redirect("contract_counterparty_list")


def process_counterparty_import_file(file_path, mode, submitter):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    mapping = {
        "\u5355\u4f4d\u540d\u79f0": "party_name",
        "\u5355\u4f4d\u7c7b\u578b": "party_type",
        "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801": "credit_code",
        "\u8054\u7cfb\u4eba": "contact_name",
        "\u8054\u7cfb\u7535\u8bdd": "contact_phone",
        "\u72b6\u6001": "status",
        "\u5907\u6ce8": "remark",
        "\u6210\u7acb\u65e5\u671f": "established_date",
        "\u6240\u5c5e\u7701\u4efd": "province_code",
        "\u6240\u5c5e\u57ce\u5e02": "city",
        "\u4f01\u4e1a\u7c7b\u578b": "enterprise_type",
        "\u6240\u5c5e\u884c\u4e1a": "industry",
        "\u66fe\u7528\u540d": "former_name",
        "\u6ce8\u518c\u5730\u5740": "registration_address",
        "\u7ecf\u8425\u8303\u56f4": "business_scope",
    }
    idx_map = {}
    for idx, header in enumerate(headers):
        if header in mapping:
            idx_map[mapping[header]] = idx

    created = updated = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {
            field: str(row[idx] if idx < len(row) and row[idx] is not None else "").strip()
            for field, idx in idx_map.items()
        }
        code = row_data.get("credit_code", "").upper()
        if not code or len(code) != 18:
            skipped += 1
            continue

        established_date = None
        if row_data.get("established_date"):
            try:
                established_date = datetime.strptime(row_data["established_date"], "%Y-%m-%d").date()
            except ValueError:
                established_date = None

        defaults = {
            "party_name": row_data.get("party_name", ""),
            "party_type": row_data.get("party_type", "OTHER_VENDOR"),
            "contact_name": row_data.get("contact_name", ""),
            "contact_phone": row_data.get("contact_phone", "")[:255],
            "status": row_data.get("status", "ACTIVE") or "ACTIVE",
            "remark": row_data.get("remark", ""),
            "established_date": established_date,
            "province_code": row_data.get("province_code", ""),
            "city": row_data.get("city", ""),
            "enterprise_type": row_data.get("enterprise_type", ""),
            "industry": row_data.get("industry", ""),
            "former_name": row_data.get("former_name", ""),
            "registration_address": row_data.get("registration_address", ""),
            "business_scope": row_data.get("business_scope", ""),
        }
        obj = Counterparty.objects.filter(credit_code=code).first()
        if obj and mode == "insert":
            skipped += 1
            continue
        if obj and mode == "upsert":
            for key, value in defaults.items():
                setattr(obj, key, value)
            obj.updated_by = submitter
            obj.save()
            updated += 1
            continue
        Counterparty.objects.create(credit_code=code, created_by=submitter, updated_by=submitter, **defaults)
        created += 1

    os.unlink(file_path)
    return {"created": created, "updated": updated, "skipped": skipped}


@login_required
def import_contract_ledger(request):
    if request.method != "POST":
        return redirect("contract_list")

    mode = request.POST.get("mode", "insert").strip().lower()
    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "\u8bf7\u9009\u62e9\u5408\u540c\u53f0\u8d26\u6587\u4ef6")
        return redirect("contract_list")

    if mode not in {"insert", "upsert"}:
        messages.error(request, "\u5bfc\u5165\u6a21\u5f0f\u4e0d\u5408\u6cd5")
        return redirect("contract_list")

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    mapping = {
        "\u9879\u76ee\u4e3b\u6570\u636e\u7f16\u7801": "project_code",
        "\u5408\u540cCT\u7801": "contract_ct_code",
        "\u5408\u540c\u540d\u79f0": "contract_name",
        "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801": "credit_code",
        "\u6765\u6e90\u7cfb\u7edf": "source_system",
        "\u5408\u540c\u65b9\u5411": "contract_direction",
        "\u5408\u540c\u5206\u7c7b": "contract_category",
        "\u5408\u540c\u5e74\u4efd": "contract_year",
        "\u539f\u59cb\u542b\u7a0e\u91d1\u989d": "original_amount_tax",
        "\u539f\u59cb\u4e0d\u542b\u7a0e\u91d1\u989d": "original_amount_notax",
        "\u539f\u59cb\u7a0e\u7387": "original_tax_rate",
        "\u5f53\u524d\u542b\u7a0e\u91d1\u989d": "current_amount_tax",
        "\u5f53\u524d\u4e0d\u542b\u7a0e\u91d1\u989d": "current_amount_notax",
        "\u5f53\u524d\u7a0e\u7387": "current_tax_rate",
        "\u5408\u540c\u72b6\u6001": "contract_status",
        "\u5907\u6ce8": "remark",
    }
    idx_map = {}
    for idx, header in enumerate(headers):
        if header in mapping:
            idx_map[mapping[header]] = idx

    for field in ["project_code", "contract_ct_code", "contract_name", "credit_code", "source_system", "contract_direction", "contract_category"]:
        if field not in idx_map:
            messages.error(request, f"\u5408\u540c\u5bfc\u5165\u7f3a\u5c11\u5fc5\u8981\u5217\uff1a{field}")
            return redirect("contract_list")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        file.seek(0)
        for chunk in file.chunks():
            tmp_file.write(chunk)
        tmp_path = tmp_file.name

    approval = _submit_business_approval(
        request,
        approval_type="import",
        target_module="contract",
        target_id=0,
        target_code=f"CT-IM-{uuid.uuid4().hex[:8].upper()}",
        target_name="\u5408\u540c\u53f0\u8d26\u5bfc\u5165",
        after_data={"mode": mode, "file_name": file.name},
        change_note=f"\u5408\u540c\u53f0\u8d26\u5bfc\u5165\u7533\u8bf7\uff08{mode}\uff09",
    )
    approval.import_file_path = tmp_path
    approval.save(update_fields=["import_file_path"])
    messages.success(request, f"\u5bfc\u5165\u5ba1\u6279\u5df2\u63d0\u4ea4\uff0c\u7b49\u5f85\u502a\u660e\u73e0\u5ba1\u6279\u3002\u5ba1\u6279\u5355\u53f7\uff1a{approval.id}")
    return redirect("contract_list")


def process_contract_import_file(file_path, mode, submitter):
    dept_name_map = _get_dept_name_map()
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    mapping = {
        "\u9879\u76ee\u4e3b\u6570\u636e\u7f16\u7801": "project_code",
        "\u5408\u540cCT\u7801": "contract_ct_code",
        "\u5408\u540c\u540d\u79f0": "contract_name",
        "\u7edf\u4e00\u793e\u4f1a\u4fe1\u7528\u4ee3\u7801": "credit_code",
        "\u6765\u6e90\u7cfb\u7edf": "source_system",
        "\u5408\u540c\u65b9\u5411": "contract_direction",
        "\u5408\u540c\u5206\u7c7b": "contract_category",
        "\u5408\u540c\u5e74\u4efd": "contract_year",
        "\u539f\u59cb\u542b\u7a0e\u91d1\u989d": "original_amount_tax",
        "\u539f\u59cb\u4e0d\u542b\u7a0e\u91d1\u989d": "original_amount_notax",
        "\u539f\u59cb\u7a0e\u7387": "original_tax_rate",
        "\u5f53\u524d\u542b\u7a0e\u91d1\u989d": "current_amount_tax",
        "\u5f53\u524d\u4e0d\u542b\u7a0e\u91d1\u989d": "current_amount_notax",
        "\u5f53\u524d\u7a0e\u7387": "current_tax_rate",
        "\u5408\u540c\u72b6\u6001": "contract_status",
        "\u5907\u6ce8": "remark",
    }
    idx_map = {}
    for idx, header in enumerate(headers):
        if header in mapping:
            idx_map[mapping[header]] = idx

    created = updated = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {
            field: str(row[idx] if idx < len(row) and row[idx] is not None else "").strip()
            for field, idx in idx_map.items()
        }
        ct_code = row_data.get("contract_ct_code", "").upper()
        if not CT_CODE_PATTERN.fullmatch(ct_code):
            skipped += 1
            continue

        project = ProjectMaster.objects.filter(project_code=row_data.get("project_code", ""), is_deleted=False).first()
        counterparty = Counterparty.objects.filter(credit_code=row_data.get("credit_code", "").upper()).first()
        if not project or not counterparty:
            skipped += 1
            continue

        try:
            defaults = {
                "project": project,
                "counterparty": counterparty,
                "contract_name": row_data.get("contract_name", ""),
                "source_system": row_data.get("source_system", "MANUAL"),
                "contract_direction": row_data.get("contract_direction", "NONE"),
                "contract_category": row_data.get("contract_category", "OTHER"),
                "undertaking_dept_code": "",
                "undertaking_dept_name": dept_name_map.get(project.dept, project.dept or ""),
                "execution_project": project,
                "execution_project_code_snapshot": project.project_code,
                "execution_project_name_snapshot": project.project_name,
                "contract_year": row_data.get("contract_year", ""),
                "original_amount_tax": _to_decimal(row_data.get("original_amount_tax", "0")),
                "original_amount_notax": _to_decimal(row_data.get("original_amount_notax", "0")),
                "original_tax_rate": _to_decimal(row_data.get("original_tax_rate"), default="0") if row_data.get("original_tax_rate") else None,
                "current_amount_tax": _to_decimal(row_data.get("current_amount_tax") or row_data.get("original_amount_tax", "0")),
                "current_amount_notax": _to_decimal(row_data.get("current_amount_notax") or row_data.get("original_amount_notax", "0")),
                "current_tax_rate": _to_decimal(row_data.get("current_tax_rate"), default="0") if row_data.get("current_tax_rate") else None,
                "contract_status": row_data.get("contract_status", "SIGNED") or "SIGNED",
                "remark": row_data.get("remark", ""),
            }
        except InvalidOperation:
            skipped += 1
            continue

        obj = ContractMaster.objects.filter(contract_ct_code=ct_code).first()
        if obj and mode == "insert":
            skipped += 1
            continue
        if obj and mode == "upsert":
            for key, value in defaults.items():
                setattr(obj, key, value)
            obj.updated_by = submitter
            obj.full_clean()
            obj.save()
            updated += 1
            continue

        new_contract = ContractMaster(contract_ct_code=ct_code, created_by=submitter, updated_by=submitter, **defaults)
        new_contract.full_clean()
        new_contract.save()
        created += 1

    os.unlink(file_path)
    return {"created": created, "updated": updated, "skipped": skipped}
